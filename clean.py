import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# columns: rc_account_id, brand, mrr, churn_date, EntryCount, dates (incremented by month 8/2019- 6/2024)
# rows: rc_account_id and their sales agreement dates


def update_new_date(df):
    """if provided a new sales agreement date, use this date to populate empty cells (if applicable) """
    new_date_column = df.columns[-1]  # column with new agreement date to update excel with
    dates_columns = df.columns[5:-2]  # date columns start at column F and end before the two rightmost columns

    for index, row in df.iterrows():  # iterate through account ids
        new_date = row[new_date_column]  # get current new date (if applicable)
        if pd.notna(new_date):  # if new date is given
            for col in reversed(dates_columns):  # traverse from right to left date columns
                if pd.isna(row[col]):  # cell is empty
                    df.at[index, col] = new_date  # populate empty cell with new date
                else:  # cell is not empty
                    break  # stop populating empty cells
    return df


def add_new_month(df):
    """adds a new column for the next month and propagates dates to the right (if applicable) """
    dates = df.columns[5:]  # date columns start at column F
    last_date_col = dates[-1]  # rightmost column (latest date in excel)
    last_date = pd.to_datetime(last_date_col, errors='coerce')   # get datetime obj
    new_month = (last_date + pd.DateOffset(months=1)).replace(day=1)  # add one month
    new_month_label = new_month.strftime('%Y-%m-%d')  # format new column for next month label
    df[new_month_label] = pd.NA  # add new column

    df = propagate_right(df)  # propagate dates to right

    return df


def same_left_right_date(df):
    """populate blank cells in between two cells that have the same sales agreement date"""
    dates = df.columns[5:]  # date columns start at column F
    for index, row in df.iterrows():  # iterate through account ids
        for date in range(len(dates) - 1):  # iterate through sale agreements for current account id
            left_date = row[dates[date]]  # current sales agreement
            if pd.notna(left_date):  # if current sales agreement is not blank
                for right_date in range(date + 1, len(dates)):  # iterate through sales agreements after current
                    if pd.notna(row[dates[right_date]]):  # found a non-blank cell
                        if left_date == row[dates[right_date]]:  # check if left and right dates are the same
                            # for the current row, populate cells between left and right date
                            for cell in range(date + 1, right_date):  # traverse through dates in between
                                df.iloc[index, data.columns.get_loc(dates[cell])] = left_date
                        break  # stop traversing once we have filled blank cells between two none blank cells
    return df


def propagate_right(df):
    """populate blank cells to the right:
    1. look for the cell in the row that has been last populated
    2. populate cells to the right with that agreement date until the
    column date surpasses it"""
    dates = df.columns[5:]  # date columns start at column F
    for index, row in df.iterrows():  # iterate through account ids
        last_populated_cell = None
        for date in range(len(dates) - 1, -1, -1):  # iterate through sale agreements from right to left
            if pd.notna(row[dates[date]]):  # if current sales agreement is not blank
                last_populated_cell = date  # last populated cell index
                break

        if last_populated_cell is not None:  # if there is a sales agreement present
            last_date = pd.to_datetime(row[dates[last_populated_cell]], errors='coerce')
            for date in range(last_populated_cell + 1,
                              len(dates)):  # iterate through dates to the right of last populated cell
                current_date = pd.to_datetime(dates[date], errors='coerce')  # column date
                if current_date <= last_date:  # populate if current column date is less than or equal to last
                    # agreement date
                    df.iloc[index, df.columns.get_loc(dates[date])] = last_date
                else:
                    break  # stop if the column date surpasses the agreement date
    return df


def choose_date(df):
    """will make a decision for what date will populate a cell:
    1. identify cells that are blank.
    2. if a cell is blank find the cell that is not blank on its left and right.
    3. identify the column date of the blank cell.
    4. if the column date's month and year are prior to the agreement date in the left cell than populate the cell
    with that agreement date
    5. otherwise populate the cell with the agreement date in the right cell"""
    dates = df.columns[5:]  # date columns start at column F
    for index, row in df.iterrows():  # iterate through account ids
        for date in range(1, len(dates) - 1):  # iterate through sale agreements for current account id
            if pd.isna(row[dates[date]]):  # current cell is blank
                left_date = None
                for left in range(date - 1, -1, -1):  # traverse cell to the left
                    if pd.notna(row[dates[left]]):  # left cell is not blank
                        left_date = pd.to_datetime(row[dates[left]], errors='coerce')  # assign date in this cell
                        break
                right_date = None
                for right in range(date + 1, len(dates)):  # traverse cells to the right
                    if pd.notna(row[dates[right]]):  # right cell is not blank
                        right_date = pd.to_datetime(row[dates[right]], errors='coerce')  # assign date in this cell
                        break
                current_column_date = pd.to_datetime(dates[date], errors='coerce')  # blank cell's column date
                if left_date is not None and right_date is not None:  # in between two cells that are populated
                    # see if column date is prior to left agreement date
                    if current_column_date.year < left_date.year:
                        df.iloc[index, df.columns.get_loc(dates[date])] = left_date  # take left date
                    elif (current_column_date.year == left_date.year) and (current_column_date.month < left_date.month):
                        df.iloc[index, df.columns.get_loc(dates[date])] = left_date  # take left date
                    else:
                        df.iloc[index, df.columns.get_loc(dates[date])] = right_date  # take right date
    return df


def needs_revision(df):
    """Identifies cells that need to be flagged for revision:
    1. sales agreement date only spans for one - four months."""
    dates = df.columns[5:]  # date columns start at column F
    to_highlight = []

    for index, row in df.iterrows():  # iterate through account ids
        for date in range(len(dates)):  # iterate through sale agreements for current account id
            current_date = row[dates[date]]  # current sales agreement being iterated
            if pd.notna(current_date):  # check that cell is not blank
                count = 1  # keep track of occurrences of current agreement date
                highlight_cells = [date]
                for right in range(date + 1, len(dates)):  # count right
                    if row[dates[right]] == current_date:  # same date
                        count += 1
                        highlight_cells.append(right)
                    else:  # stop counting
                        break
                for left in range(date - 1, -1, -1):  # count left
                    if row[dates[left]] == current_date:  # same date
                        count += 1
                        highlight_cells.insert(0, left)
                    else:  # stop counting
                        break
                if 1 <= count <= 4:  # sales agreement date only spans for 1-4 months
                    df.at[index, 'Status'] = '?'  # flag row
                    for cell in highlight_cells:
                        to_highlight.append((index + 2, df.columns.get_loc(dates[cell]) + 1))
    return df, to_highlight


def mark_done(df):
    """flag row as done if cells are populated properly:
    1. find the first cell populated in the row.
    2. find the last cell populated in the row.
    3. check for any blank cells between the first and last cells populated.
    4. if no blanks are found, mark it as done"""
    dates = df.columns[5:]  # date columns start at column F
    for index, row in df.iterrows():  # iterate through account ids
        first_populated_cell = None
        last_populated_cell = None

        # Find the first populated cell
        for date in range(len(dates)):
            if pd.notna(row[dates[date]]):
                first_populated_cell = date
                break

        # Find the last populated cell
        for date in range(len(dates) - 1, -1, -1):
            if pd.notna(row[dates[date]]):
                last_populated_cell = date
                break

        # Check for blank cells between the first and last populated cells
        if first_populated_cell is not None and last_populated_cell is not None:
            if all(pd.notna(row[dates[first_populated_cell:last_populated_cell + 1]])):
                df.at[index, 'Status'] = 'Done'
    return df


def fix_date_format(df):
    """changes cell dates and column names from datetime objs to formatted dates"""
    date_columns = data.columns[5:]  # define columns that represent dates
    for col in date_columns:  # iterate through date column cells
        df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%-m/%-d/%Y')
    # change column names to formatted dates
    formatted_date_columns = pd.to_datetime(date_columns, errors='coerce').strftime('%-m/%-d/%Y')
    df.rename(columns=dict(zip(date_columns, formatted_date_columns)), inplace=True)
    return df


# calling functions for data clean up
file_input = 'final2.xlsx'  # define excel export
data = pd.read_excel(file_input)  # open and read excel

# data = add_new_month(data)  # add new month column and propagate right accordingly
# data = same_left_right_date(data)  # calling clean up
# data = propagate_right(data)  # propagate dates to the right
# data = choose_date(data)  # populate blank cells in between different agreement dates
# data = mark_done(data)
# data = update_new_date(data)  # use new sales agreement date if applicable
data = fix_date_format(data)  # fix formatting
data, cells_to_highlight = needs_revision(data)  # clean data and add flags

output_file = 'highlighted.xlsx'  # write to this output file
data.to_excel(output_file, index=False)  # writing updated data frame to output file

# Apply conditional formatting to highlight cells that need revision
wb = openpyxl.load_workbook(output_file)
ws = wb.active
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Iterate through the list of cells to apply formatting
for (r, c) in cells_to_highlight:
    ws.cell(row=r, column=c).fill = yellow_fill

# Save the workbook
wb.save(output_file)

print("finished clean up")
